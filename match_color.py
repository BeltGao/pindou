import pandas as pd
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os


# ==============================
def read_color_csv(file_path):
    """尝试多种编码读取CSV，直到成功"""
    encodings = ['utf-8', 'gbk', 'gb2312', 'big5', 'latin1', 'cp1252']
    for enc in encodings:
        try:
            return pd.read_csv(file_path, encoding=enc)
        except UnicodeDecodeError:
            continue
    raise ValueError('无法识别文件编码，请将CSV文件另存为UTF-8编码')
def color_match_pindou(csv_path,image_path,output_path,cell_size_px):
    # 1. 读取色号CSV
    # 读取色号表（支持CSV和Excel）
    ext = os.path.splitext(csv_path)[1].lower()
    if ext == '.csv':
        df = read_color_csv(csv_path)
    elif ext in ('.xlsx', '.xls'):
        df = pd.read_excel(csv_path)
    else:
        raise ValueError('色号表文件格式不支持，请使用CSV或Excel文件')

    df.columns = [str(c).strip() for c in df.columns]

    required = {'色号名', 'R', 'G', 'B'}
    if not required.issubset(df.columns):
        raise ValueError(f'CSV缺少必要列：{required - set(df.columns)}，实际列名：{list(df.columns)}')

    # 构建色号列表
    color_list = []
    for _, row in df.iterrows():
        try:
            r = int(row['R'])
            g = int(row['G'])
            b = int(row['B'])
            name = str(row['色号名']).strip()
            color_list.append((r, g, b, name))
        except Exception:
            continue

    print(f'已加载 {len(color_list)} 个色号')

    # 2. 读取图片
    img = Image.open(image_path).convert('RGB')
    width, height = img.size
    pixels = img.load()
    print(f'图片尺寸：{width} x {height}')

    # Excel行列限制检查
    MAX_COL = 16384
    MAX_ROW = 1048576
    if width > MAX_COL or height > MAX_ROW:
        raise ValueError(f'图片尺寸 {width}x{height} 超过Excel限制（{MAX_COL}列 x {MAX_ROW}行），请先缩小图片。')

    # 3. 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = '像素色号'

    # 设置列宽和行高，使单元格近似正方形
    col_width = cell_size_px / 7.5
    row_height = cell_size_px * 0.75
    for c in range(1, width + 1):
        ws.column_dimensions[get_column_letter(c)].width = col_width
    for r in range(1, height + 1):
        ws.row_dimensions[r].height = row_height

    # 填充缓存与对齐
    fill_cache = {}
    center_alignment = Alignment(horizontal='center', vertical='center')

    # 4. 查找最近色号的函数（返回色号名和对应的RGB）
    def closest_color(r, g, b):
        min_dist = float('inf')
        best_name = ''
        best_rgb = (0, 0, 0)
        for cr, cg, cb, name in color_list:
            dist = (r - cr) ** 2 + (g - cg) ** 2 + (b - cb) ** 2
            if dist < min_dist:
                min_dist = dist
                best_name = name
                best_rgb = (cr, cg, cb)
                if dist == 0:  # 完全匹配，提前结束
                    break
        return best_name, best_rgb

    # 5. 遍历像素并写入Excel
    for y in range(height):
        for x in range(width):
            r, g, b = pixels[x, y]
            name, matched_rgb = closest_color(r, g, b)  # 获取匹配到的色号名和颜色
            mr, mg, mb = matched_rgb

            cell = ws.cell(row=y + 1, column=x + 1)
            cell.value = name
            cell.alignment = center_alignment

            # 用匹配到的色号颜色填充单元格（而不是像素原色）
            hex_color = f'FF{mr:02X}{mg:02X}{mb:02X}'
            if hex_color not in fill_cache:
                fill_cache[hex_color] = PatternFill(start_color=hex_color,
                                                    end_color=hex_color,
                                                    fill_type='solid')
            cell.fill = fill_cache[hex_color]

        # 进度提示
        if (y + 1) % 10 == 0 or (y + 1) == height:
            print(f'已处理 {y + 1}/{height} 行')

    # 6. 保存
    wb.save(output_path)
    print(f'完成！已保存到：{output_path}')
if __name__ == '__main__':
    # ========== 参数设置 ==========
    csv_path = '色号.csv'  # 色号表路径
    image_path = '3.png'  # 图片路径
    output_path = '输出2.xlsx'  # 输出Excel路径
    cell_size_px = 30  # 单元格近似边长（像素）
    color_match_pindou( csv_path,image_path,output_path,cell_size_px)

